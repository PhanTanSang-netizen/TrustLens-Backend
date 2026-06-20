from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="EEF5FF")


def _safe_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def _write_key_value_rows(
    sheet,
    start_row: int,
    title: str,
    data: dict[str, Any],
) -> int:
    sheet.cell(row=start_row, column=1, value=title)
    sheet.cell(row=start_row, column=1).font = Font(bold=True, size=12)
    sheet.cell(row=start_row, column=1).fill = SECTION_FILL

    row = start_row + 1

    for key, value in data.items():
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=_safe_value(value))
        sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    return row + 1


def _style_header_row(sheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_fit_columns(sheet, max_width: int = 45) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def _style_body(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_summary_sheet(
    workbook: Workbook,
    report_data: dict[str, Any],
) -> None:
    sheet = workbook.active
    sheet.title = "Summary"

    sheet["A1"] = "TrustLens Report Summary"
    sheet["A1"].font = Font(bold=True, size=16)

    submission = report_data.get("submission") or {}
    file_info = report_data.get("file") or {}
    summary = report_data.get("summary") or {}
    processing = summary.get("processing") or {}
    verification = summary.get("verification") or {}
    score = summary.get("score") or {}

    row = 3

    row = _write_key_value_rows(
        sheet=sheet,
        start_row=row,
        title="Submission",
        data={
            "Submission ID": submission.get("id"),
            "Assignment ID": submission.get("assignment_id"),
            "Owner Label": submission.get("owner_label"),
            "Status": submission.get("status"),
            "Overall Score": submission.get("overall_score"),
            "Created At": submission.get("created_at"),
        },
    )

    row = _write_key_value_rows(
        sheet=sheet,
        start_row=row,
        title="File",
        data={
            "Original Name": file_info.get("original_name"),
            "MIME Type": file_info.get("mime_type"),
            "Size Bytes": file_info.get("size_bytes"),
            "Checksum": file_info.get("checksum"),
        },
    )

    row = _write_key_value_rows(
        sheet=sheet,
        start_row=row,
        title="Processing",
        data={
            "Submission Status": processing.get("submission_status"),
            "Has Extracted Text": processing.get("has_extracted_text"),
            "Has Reference Section": processing.get("has_reference_section"),
            "Citation Count": processing.get("citation_count"),
            "Metadata Record Count": processing.get("metadata_record_count"),
        },
    )

    row = _write_key_value_rows(
        sheet=sheet,
        start_row=row,
        title="Verification",
        data={
            "Total": verification.get("total"),
            "Verified": verification.get("verified"),
            "Basic Metadata Present": verification.get("basic_metadata_present"),
            "Broken": verification.get("broken"),
            "Forbidden": verification.get("forbidden"),
            "Unreachable": verification.get("unreachable"),
            "Not Provided": verification.get("not_provided"),
        },
    )

    _write_key_value_rows(
        sheet=sheet,
        start_row=row,
        title="Score",
        data={
            "Overall Score": score.get("overall_score"),
            "Trust Level": score.get("trust_level"),
            "Note": score.get("note"),
        },
    )

    _style_body(sheet)
    _auto_fit_columns(sheet)


def _add_citations_sheet(
    workbook: Workbook,
    report_data: dict[str, Any],
) -> None:
    sheet = workbook.create_sheet("Citations")

    headers = [
        "Sequence",
        "Detected Style",
        "Authors",
        "Title",
        "Year",
        "DOI",
        "URL",
        "Verification Status",
        "Provider",
        "Confidence Score",
        "Trust Level",
        "Score",
        "Warnings",
        "Raw Text",
    ]

    sheet.append(headers)
    _style_header_row(sheet, row=1, max_col=len(headers))

    citations = report_data.get("citations") or []

    for item in citations:
        citation = item.get("citation") or {}
        metadata = item.get("metadata") or {}
        warnings = item.get("warnings") or []

        sheet.append(
            [
                _safe_value(citation.get("sequence_no")),
                _safe_value(citation.get("detected_style")),
                _safe_value(citation.get("authors")),
                _safe_value(citation.get("title")),
                _safe_value(citation.get("year")),
                _safe_value(citation.get("doi")),
                _safe_value(citation.get("url")),
                _safe_value(metadata.get("verification_status")),
                _safe_value(metadata.get("provider")),
                _safe_value(metadata.get("confidence_score")),
                _safe_value(item.get("trust_level")),
                _safe_value(item.get("score")),
                ", ".join(warnings) if warnings else "",
                _safe_value(citation.get("raw_text")),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    _style_body(sheet)
    _auto_fit_columns(sheet)


def _add_metadata_sheet(
    workbook: Workbook,
    report_data: dict[str, Any],
) -> None:
    sheet = workbook.create_sheet("Metadata")

    headers = [
        "Citation Sequence",
        "Provider",
        "Query Type",
        "Query Value",
        "Source URL",
        "Matched Title",
        "Matched Year",
        "Verification Status",
        "Confidence Score",
    ]

    sheet.append(headers)
    _style_header_row(sheet, row=1, max_col=len(headers))

    citations = report_data.get("citations") or []

    for item in citations:
        citation = item.get("citation") or {}
        metadata = item.get("metadata") or {}

        sheet.append(
            [
                _safe_value(citation.get("sequence_no")),
                _safe_value(metadata.get("provider")),
                _safe_value(metadata.get("query_type")),
                _safe_value(metadata.get("query_value")),
                _safe_value(metadata.get("source_url")),
                _safe_value(metadata.get("matched_title")),
                _safe_value(metadata.get("matched_year")),
                _safe_value(metadata.get("verification_status")),
                _safe_value(metadata.get("confidence_score")),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    _style_body(sheet)
    _auto_fit_columns(sheet)


def build_xlsx_report_bytes(
    report_data: dict[str, Any],
) -> bytes:
    workbook = Workbook()

    _add_summary_sheet(
        workbook=workbook,
        report_data=report_data,
    )

    _add_citations_sheet(
        workbook=workbook,
        report_data=report_data,
    )

    _add_metadata_sheet(
        workbook=workbook,
        report_data=report_data,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer.read()


def export_report_to_xlsx_bytes(
    report_data: dict[str, Any],
) -> bytes:
    return build_xlsx_report_bytes(report_data)


def export_report_to_xlsx_file(
    report_data: dict[str, Any],
    output_path: str | Path,
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(build_xlsx_report_bytes(report_data))
    return str(output_path)


def generate_xlsx_report_filename(
    submission_id: str,
) -> str:
    return f"trustlens_report_{submission_id}_{uuid4().hex[:8]}.xlsx"